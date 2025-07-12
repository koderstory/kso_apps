import base64
import csv
import io

from odoo import models, fields, api, _
from odoo.exceptions import UserError

class CategoryImportWizard(models.TransientModel):
    _name = "category.import.wizard"
    _description = "Category Import Wizard"

    file = fields.Binary(string="File", required=True)
    filename = fields.Char(string="Filename")

    def action_import(self):
        self.ensure_one()
        try:
            if not self.file:
                raise UserError(_("No file provided for import."))

            # Create a mutable counter to track the number of inserted categories.
            count_container = {"inserted_count": 0}
            # Create a cache dictionary to store created/found categories by full hierarchical path.
            category_cache = {}

            # Determine the file type from the filename extension.
            ext = self.filename.split('.')[-1].lower() if self.filename else 'csv'
            
            if ext in ('csv',):
                # Process as CSV.
                file_data = base64.b64decode(self.file)
                try:
                    data_str = file_data.decode('utf-8')
                except UnicodeDecodeError:
                    data_str = file_data.decode('latin1')
                file_io = io.StringIO(data_str, newline='')
                csv_reader = csv.DictReader(file_io)
                for row in csv_reader:
                    # Expecting a column header 'name' for the hierarchical path.
                    path = row.get('name', '').strip()
                    if path:
                        # Optionally, get the 'code' if it exists.
                        code = row.get('code', None)
                        if code:
                            code = code.strip()
                        self._create_category_hierarchy(path, category_cache, code, count_container)
            elif ext in ('xls', 'xlsx'):
                # Process as Excel using openpyxl.
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    raise UserError(_("The openpyxl module is not installed. Please install it to import Excel files."))
                file_data = base64.b64decode(self.file)
                workbook = load_workbook(filename=io.BytesIO(file_data), read_only=True)
                sheet = workbook.active
                headers = []
                name_index = None
                code_index = None
                for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if i == 1:
                        headers = [str(cell).strip() if cell else "" for cell in row]
                        try:
                            name_index = headers.index("name")
                        except ValueError:
                            raise UserError(_("Excel file must have a 'name' column in the header."))
                        if "code" in headers:
                            code_index = headers.index("code")
                    else:
                        if not row:
                            continue
                        category_path = row[name_index]
                        if category_path:
                            category_path = str(category_path).strip()
                            code = None
                            if code_index is not None:
                                code_val = row[code_index]
                                if code_val:
                                    code = str(code_val).strip()
                            self._create_category_hierarchy(category_path, category_cache, code, count_container)
            else:
                raise UserError(_("Unsupported file format. Please upload a CSV or Excel file."))
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Success'),
                    'message': _('Product categories imported successfully. %s categories inserted.') % count_container["inserted_count"],
                    'type': 'success',
                    'sticky': False,
                    'next': {'type': 'ir.actions.act_window_close'},
                },
            }
        except Exception as e:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Error'),
                    'message': _('Failed to import product categories. Error: %s') % str(e),
                    'type': 'danger',
                    'sticky': False,
                    'next': {'type': 'ir.actions.act_window_close'},
                },
            }

    def _create_category_hierarchy(self, path, cache, code=None, counter=None):
        """Creates or retrieves category hierarchy from a path string.
        Example: "material / supporting / screw"

        :param path: The hierarchical path string.
        :param cache: A dictionary to cache categories by their full hierarchical path.
        :param code: Optional code to assign to the leaf category.
        :param counter: A dictionary with a key 'inserted_count' to track new creations.
        :return: The leaf product.category record.
        """
        Category = self.env["product.category"]
        # Check if the model has a 'code' field.
        has_code_field = 'code' in Category._fields

        parts = [p.strip() for p in path.split("/") if p.strip()]
        full_path = ""
        parent = None
        for i, part in enumerate(parts):
            full_path = full_path + ("/" if full_path else "") + part
            if full_path in cache:
                parent = cache[full_path]
                # If this is the leaf category and a code is provided, update it if applicable.
                if has_code_field and i == len(parts) - 1 and code:
                    parent.write({'code': code})
                continue

            domain = [("name", "=", part)]
            if parent:
                domain.append(("parent_id", "=", parent.id))
            else:
                domain.append(("parent_id", "=", False))

            category = Category.search(domain, limit=1)
            if not category:
                values = {
                    "name": part,
                    "parent_id": parent.id if parent else False,
                }
                if has_code_field and i == len(parts) - 1 and code:
                    values["code"] = code
                category = Category.create(values)
                if counter is not None:
                    counter["inserted_count"] += 1
            else:
                if has_code_field and i == len(parts) - 1 and code:
                    category.write({"code": code})
            cache[full_path] = category
            parent = category
        return parent
