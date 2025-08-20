# -*- coding: utf-8 -*-
"""
This module defines the Product Variant Import Wizard.
It provides a wizard form where the user can upload an Excel file
containing product and variant data, and includes a button to download
a sample template file.
"""

import base64
import io
import openpyxl
import logging

from odoo import models, fields, api, _
from odoo.exceptions import UserError

# Import the main helper function from the helper file.
from .import_variant_helpers import add_or_update_product_with_variants

_logger = logging.getLogger(__name__)


class ProductVariantImportWizard(models.TransientModel):
    """
    Transient Model for importing product variants from an Excel file.
    
    This wizard lets the user select an Excel file, which is then processed
    to update or create product templates and their associated variants.
    It also includes functionality to download a sample/template file.
    """
    _name = "kso.import.productvariant.wizard"
    _description = "Product Variant Import Wizard"

    file = fields.Binary(string="File", required=True)
    filename = fields.Char(string="Filename")

    def action_import_productvariant(self):
        """
        Import product variants from the provided Excel file (streaming + bounded).
        """
        self.ensure_one()

        if not self.file:
            raise UserError(_("Please provide a file to import."))

        # 1) Decode
        try:
            file_data = base64.b64decode(self.file)
        except Exception:
            raise UserError(_("The file could not be decoded. Please try again."))

        # 2) Open in READ-ONLY streaming mode
        try:
            workbook = openpyxl.load_workbook(
                io.BytesIO(file_data),
                read_only=True,      # <- critical: stream rows
                data_only=True       # <- use cached formula results if any
            )
            sheet = workbook.active
        except Exception as e:
            raise UserError(_("Error reading Excel file: %s") % str(e))

        # 3) Determine actual used range and build header safely
        from openpyxl.utils.cell import range_boundaries
        dim = sheet.calculate_dimension()  # e.g. "A1:K2045"
        min_col, min_row, max_col, max_row = range_boundaries(dim)

        # Stream the header row (no sheet[1] — that materializes cells)
        header_row_iter = sheet.iter_rows(
            min_row=1, max_row=1,
            min_col=min_col, max_col=max_col,
            values_only=True
        )
        try:
            header_row = next(header_row_iter)
        except StopIteration:
            raise UserError(_("The sheet appears to be empty."))

        # Normalize headers
        headers = [(str(h).strip().lower() if h is not None else '') for h in header_row]
        # Expected columns (you can tweak this list)
        wanted = [
            "name", "product code", "uom", "purchase uom", "type",
            "is tracked", "tracked by", "category", "sale price",
            "cost price", "stock quantity", "supplier", "variant"
        ]
        # Map header -> column index within the streamed window
        idx = {h: i for i, h in enumerate(headers) if h}

        # Bound the max_col to only the last needed header we actually found
        needed_ix = [idx[h] for h in wanted if h in idx]
        if not needed_ix:
            raise UserError(_("No expected headers were found in the file: %s") % ", ".join(wanted))
        bound_max_col = min_col + max(needed_ix)

        # 4) Stream data rows, only across the needed rectangle
        product_data = []
        for row in sheet.iter_rows(
            min_row=2, max_row=max_row,
            min_col=min_col, max_col=bound_max_col,
            values_only=True
        ):
            rec = {}
            # Build dict with only the columns we use
            for h in wanted:
                j = idx.get(h)
                if j is not None and j < len(row):
                    rec[h] = row[j]
            # Only add non-empty rows (at least one significant field)
            if any(rec.get(k) not in (None, "", 0) for k in ("name", "variant", "product code", "category")):
                product_data.append(rec)

        # Optional: avoid logging the whole dataset (could be huge)
        _logger.info("Product import rows parsed: %s", len(product_data))

        # 5) Process with your helper
        try:
            add_or_update_product_with_variants(self.env, product_data)
        except Exception as e:
            # Bubble up as a user-visible message
            raise UserError(_("Import failed: %s") % e)

        return {'type': 'ir.actions.act_window_close'}


    def action_download_template(self):
        """
        Provide a download link for the sample product variant Excel template.
        
        This method returns an action that directs the user to the URL of
        the static template file. Ensure the file is located at:
            static/description/DEMO_DATA.xlsx
        
        :return: An action of type 'ir.actions.act_url' to download the template.
        """
        # Retrieve the base URL configured in Odoo.
        base_url = self.env['ir.config_parameter'].sudo().get_param('web.base.url')
        return {
            'type': 'ir.actions.act_url',
            # Concatenate the base URL with your relative static file path.
            'url': base_url + '/kso_import_productvariant/static/description/DEMO_DATA.xlsx?download=true',
            'target': 'self',
        }
